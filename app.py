import sys
import os
import time
import tempfile
from pathlib import Path
from datetime import datetime,date
from typing import Generator
from decimal import Decimal
import sqlite3
# import pyodbc

# PyQt5
from PyQt5.QtCore import QDate
from PyQt5.QtGui import QColor, QPixmap
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QMessageBox, QFileDialog, QTableWidgetItem
)
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QCompleter
from PyQt5.QtGui import QDoubleValidator,QIntValidator
from PyQt5.uic import loadUiType
# from PyQt5.QtWebEngineWidgets import QWebEngineView

# SQL & Database
import pyodbc
from sqlalchemy import create_engine, VARCHAR, NVARCHAR, INTEGER, DATE, DECIMAL
from sqlalchemy.engine import URL, Engine
from sqlalchemy.orm import sessionmaker, declarative_base

# Excel & Pandas
import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
# Load UI
import qdarkstyle
import resources_rc

def get_resource_path(relative_path):
    """Trả về đường dẫn đầy đủ đến tài nguyên."""
    if getattr(sys, 'frozen', False):  # Kiểm tra nếu đang chạy file .exe
        base_path = sys._MEIPASS
    else:  # Nếu đang chạy bằng Python gốc
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def get_app_dir():
    # Nếu đã đóng gói bằng PyInstaller
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)  # đường dẫn thư mục chứa file .exe
    else:
        return os.path.dirname(os.path.abspath(__file__))  # đường dẫn file .py đang chạy

def get_config_path():
    app_dir = get_app_dir()
    config_path = os.path.join(app_dir, "config_link.txt")
    
    if not os.path.exists(config_path):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle("Lỗi cấu hình")
        msg.setText("❌ Không tìm thấy file cấu hình 'config_link.txt'!")
        msg.exec_()
        return None

    return config_path

def read_db_path():
    config_path = get_config_path()
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return f.readline().strip()
    except FileNotFoundError:
        return None

def connect_to_db(): 
    try:
        db_file = read_db_path()
        try:
            connection = sqlite3.connect(db_file)       
            return connection
        except:
            return None
    except sqlite3.Error as e:
        return None
import smtplib
from email.mime.text import MIMEText

def send_email(to_email, subject, message):
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "your_email@gmail.com"
    sender_password = "app_password"  # Không dùng mật khẩu Gmail thật

    msg = MIMEText(message)
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = to_email

    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(sender_email, sender_password)
    server.send_message(msg)
    server.quit()

def table_to_dataframe(table_widget,headers):
        rows = table_widget.rowCount()
        columns = table_widget.columnCount()
        
        # Lấy tiêu đề cột
        # headers = [table_widget.horizontalHeaderItem(i).text() for i in range(columns)]
        
        # Lấy dữ liệu từ bảng
        data = []
        for row in range(rows):
            row_data = []
            for column in range(columns):
                item = table_widget.item(row, column)
                row_data.append(item.text() if item else '')  # Lấy text từ ô, nếu không có thì gán chuỗi rỗng
            data.append(row_data)
        
        # Tạo DataFrame
        df = pd.DataFrame(data, columns=headers)
        return df
    
ui, _ = loadUiType(get_resource_path('app.ui'))

class MainApp(QMainWindow,ui):
    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        
        self.tabWidget.setCurrentIndex(0)
        self.tabWidget.tabBar().setVisible(False)
        self.menuBar.setVisible(False)
        self.bt000.clicked.connect(self.show_notifications)
        self.bt000.setVisible(False)

        # self.toolBar.setVisible(False)
        self.bt001.clicked.connect(self.login)
        self.bt203.clicked.connect(self.insert_cong_viec)
        self.bt300.clicked.connect(self.insert_chi_tiet_cong_viec)
        self.bt204.clicked.connect(self.save_cong_viec)
        self.bt301.clicked.connect(self.save_chi_tiet_cong_viec)
        self.bt205.clicked.connect(self.delete_cong_viec)
        self.bt302.clicked.connect(self.delete_chi_tiet_cong_viec)
        self.bt401.clicked.connect(self.change_pw)
        self.bt501.clicked.connect(self.add_chuc_danh)
        self.bt502.clicked.connect(self.insert_CBCNV)
        self.bt503.clicked.connect(self.save_CBCNV)
        self.bt504.clicked.connect(self.delete_CBCNV)
        self.menu11.triggered.connect(self.show_tab_1)
        self.menu13.triggered.connect(self.show_tab_2)
        self.menu15.triggered.connect(self.show_tab_5)
        self.menu51.triggered.connect(self.show_login_tab)
        self.menu52.triggered.connect(self.show_tab_4)
        self.menu61.triggered.connect(self.tai_xuong_danh_sach_cong_viec)
        self.menu62.triggered.connect(self.tai_xuong_file_excel)

        self.de101.setDate(QDate.currentDate().addDays(-365))
        self.de101.setCalendarPopup(True)
        self.de102.setDate(QDate.currentDate())
        self.de102.setCalendarPopup(True)
        self.de301.setDate(QDate.currentDate())
        self.de301.setCalendarPopup(True)  # Bật popup chọn lịch
        self.de302.setDate(QDate.currentDate())  # Gán ngày hiện tại
        self.de302.setCalendarPopup(True)

        self.le304.setValidator(QDoubleValidator(0.1, 100, 2))
        self.le305.setValidator(QIntValidator(0,100))
        self.le306.setValidator(QIntValidator(1,5))
        self.le307.setValidator(QIntValidator(1,5))
        # self.bt204.clicked.connect(self.tai_xuong_file_mau_Checker)
        # ####
        self.cb200.currentIndexChanged.connect(self.change_cb200)
        self.cb301.currentIndexChanged.connect(self.change_cb301)
        self.cb500.currentIndexChanged.connect(self.change_cb500)
        self.cb501.currentIndexChanged.connect(self.load_lb501)
        self.cb501.currentIndexChanged.connect(self.load_cb502)
        self.cb502.currentIndexChanged.connect(self.load_cb503)
        self.cb202.currentIndexChanged.connect(self.load_lb202)
        self.cb203.currentIndexChanged.connect(self.load_cb204)
        self.cb203.currentIndexChanged.connect(self.load_cb202)
        self.cb204.currentIndexChanged.connect(self.load_lb204)
        self.cb204.currentIndexChanged.connect(self.load_cb205)
        self.cb205.currentIndexChanged.connect(self.load_lb205)
        self.cb205.currentIndexChanged.connect(self.load_cb206)
        self.cb206.currentIndexChanged.connect(self.load_lb206)
        self.cb304.currentIndexChanged.connect(self.load_lb304)
        self.bt201.clicked.connect(self.add_du_an)
        self.bt207.clicked.connect(self.add_cong_viec)
        self.bt303.clicked.connect(self.add_chi_tiet_cong_viec)
        self.bt304.clicked.connect(self.show_tab_1)
        self.rd201.toggled.connect(self.rd201_change)
        self.rd301.toggled.connect(self.rd301_change)
        self.rd501.toggled.connect(self.rd501_change)
        self.bt102.clicked.connect(self.search_cong_viec)
        self.tableWidget.itemDoubleClicked.connect(self.on_table_double_click)

    def check_notifications(self):
        connection = connect_to_db()
        if connection is None:
            return
        cursor = connection.cursor()

        # Việc mới giao hôm nay
        cursor.execute("""
            SELECT COUNT(*) FROM CHI_TIET_CONG_VIEC 
            WHERE date(Ngay_bat_dau) = date('now')
                       AND Trang_thai != 'Đã hoàn thành'
        """)
        new_jobs = cursor.fetchone()[0]

        # Việc sắp hết hạn trong 2 ngày và chưa hoàn thành
        cursor.execute("""
            SELECT COUNT(*) FROM CHI_TIET_CONG_VIEC 
            WHERE date(Thoi_han) <= date('now', '+2 day')
            AND Trang_thai != 'Đã hoàn thành'
        """)
        due_soon = cursor.fetchone()[0]

        total_notify = new_jobs + due_soon
        if total_notify > 0:
            self.bt000.setText(f"🔔 {total_notify}")
        else:
            self.bt000.setText("🔔")
    def show_notifications(self):
        connection = connect_to_db()
        mnv = self.lb004.text() + '-' + self.lb001.text()
        phan_quyen = self.lb005.text()
        ma_pb = self.lb000.text()
        if connection is None:
            self.msgbox("Không thể kết nối CSDL!")
            return
        cursor = connection.cursor()
        if phan_quyen == "Nhân viên":
            cursor.execute("""
                SELECT Chi_tiet_cong_viec, Thoi_han, Tien_do,ID_CV,ID,Nguoi_thuc_hien
                FROM CHI_TIET_CONG_VIEC 
                WHERE 
                    (date(Ngay_bat_dau) = date('now')
                    OR (date(Thoi_han) <= date('now', '+2 day') AND Trang_thai != 'Đã hoàn thành'))
                    AND Nguoi_thuc_hien = ?
                ORDER BY Thoi_han ASC
            """,(mnv,))
        else:
            cursor.execute("""
                SELECT Chi_tiet_cong_viec, Thoi_han, Tien_do,ID_CV,ID,Nguoi_thuc_hien
                FROM CHI_TIET_CONG_VIEC 
                WHERE 
                    (date(Ngay_bat_dau) = date('now')
                    OR date(Thoi_han) <= date('now', '+2 day'))
                    AND ID_CV IN (SELECT ID_CV FROM DANH_SACH_CONG_VIEC 
                           WHERE Ma_phong_ban = ?)
                    AND Trang_thai != 'Đã hoàn thành'
                ORDER BY Thoi_han ASC
            """,(ma_pb,))
        data = cursor.fetchall()

        if not data:
            self.msgbox("Không có thông báo công việc.")
            return

        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel

        dlg = QDialog(self)
        dlg.setWindowTitle("Thông báo công việc")
        layout = QVBoxLayout()
        for item in data:
            line = f"- {item[0]} (Hạn: {item[1]}, Tiến độ: {int(item[2])}%, ID công việc: {item[3]}, ID chi tiết công việc: {item[4]}, Người thực hiện: {item[5]})"
            layout.addWidget(QLabel(line))
        dlg.setLayout(layout)
        dlg.exec_()

    def rd201_change(self):
        if self.rd201.isChecked():
            self.cb200.clear()
            self.bt203.setVisible(True)
            self.bt204.setVisible(False)
            self.bt205.setVisible(False)
            self.cb200.setVisible(False)
            self.lb200.setVisible(False)
            self.lb003.setText("Thêm công việc mới")
        else:
            self.load_cb200()
            self.bt203.setVisible(False)
            self.bt204.setVisible(True)
            self.bt205.setVisible(True)
            self.cb200.setVisible(True)
            self.lb200.setVisible(True)
            self.lb003.setText("Sửa hoặc xóa công việc")
    def rd301_change(self):
        if self.rd301.isChecked():
            self.cb301.clear()
            self.bt300.setVisible(True)
            self.bt301.setVisible(False)
            self.bt302.setVisible(False)
            self.cb301.setVisible(False)
            self.lb301.setVisible(False)
            self.widget_302.setVisible(False)
        else:
            # self.load_cb200()
            self.bt300.setVisible(False)
            self.bt301.setVisible(True)
            self.bt302.setVisible(True)
            self.cb301.setVisible(True)
            self.lb301.setVisible(True)
            if self.lb005.text() == "Quản lý":
                self.widget_302.setVisible(True)
            else:
                self.widget_302.setVisible(False)
            self.load_cb301()
    def rd501_change(self):
        if self.rd501.isChecked():
            self.le501.setText("")
            self.le502.setText("")
            self.le503.setText("")
            self.bt502.setVisible(True)
            self.bt503.setVisible(False)
            self.bt504.setVisible(False)
            self.widget_501.setVisible(False)
        else:
            self.widget_501.setVisible(True)
            self.bt502.setVisible(False)
            self.bt503.setVisible(True)
            self.bt504.setVisible(True)
            self.load_cb500()
            self.load_cb501()
    def msgbox(self, message):
        QtWidgets.QMessageBox.information(self, "Thông báo", message)

    def login(self):
        # fty = self.cb001.currentText()
        un = self.tb001.text()
        pw = self.tb002.text()
        
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            self.msgbox("❌Lỗi khi kết nối tới cơ sở dữ liệu")
            return
        cursor = connection.cursor()
        cursor.execute("""
                SELECT Ma_phong_ban, MNV, Ho_ten , Phan_quyen,Chuc_danh
                FROM DANH_SACH_CBCNV 
                WHERE MNV = ? AND Mat_khau = ?
            """, (un, pw))
        result = cursor.fetchone()
        if connection:
            connection.close()    

        if result:
            self.menuBar.setVisible(True)
            self.tabWidget.setCurrentIndex(1)
            self.lb003.setText("Danh sách công việc")
            self.lb000.setText(result[0])
            self.lb004.setText(result[1])
            self.lb001.setText(result[2])
            self.lb005.setText(result[3])
            self.lb401.setText(result[1])
            self.lb402.setText(result[4])
            self.cb104.setCurrentIndex(self.cb104.findText(result[0]))

            self.search_cong_viec()
            self.bt000.setVisible(True)
            self.check_notifications()
        else:
            self.lb002.setText("Tài khoản hoặc mật khẩu không đúng!")
    def change_pw(self):
        un = self.lb004.text()
        pw = self.le401.text()
        new_pw = self.le402.text().strip()
        cf_new_pw = self.le403.text().strip()

        if new_pw != cf_new_pw:
            self.msgbox("⚠️ Mật khẩu mới phải trùng với xác nhận mật khẩu mới")
            return
        
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        cursor.execute("""
                SELECT Ma_phong_ban, MNV, Ho_ten , Phan_quyen
                FROM DANH_SACH_CBCNV 
                WHERE MNV = ? AND Mat_khau = ?
            """, (un, pw))
        result = cursor.fetchone()   

        if result:
            cursor.execute("""
                UPDATE DANH_SACH_CBCNV 
                SET Mat_khau = ?
                WHERE MNV = ?
            """, (new_pw, un))
            connection.commit()
            self.msgbox("✅ Đổi mật khẩu thành công")
            self.le401.setText("")
            self.le402.setText("")
            self.le403.setText("")
            self.show_login_tab()
        else:
            self.msgbox("⚠️ Mật khẩu hiện tại không đúng")
            return
        if connection:
            connection.close() 
    def add_du_an(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_du_an = self.le201.text().strip()
        ten_du_an = self.le202.text().strip()
        # ten_du_an = self.te201.toPlainText()

        if ma_du_an and ten_du_an:
            try:
                cursor.execute(
                    "INSERT INTO DANH_SACH_DU_AN (ma_du_an, ten_du_an) VALUES (?, ?)",
                    (ma_du_an, ten_du_an)
                )
                connection.commit()
                self.msgbox("✅ Thêm dự án thành công")
            except Exception as e:
                print("Lỗi khi thêm dự án:", e)
                self.msgbox("⚠️ Mã dự án đã tồn tại hoặc có lỗi khác")
        else:
            self.msgbox("⚠️ Vui lòng nhập đầy đủ mã và tên dự án")

        connection.close()
        self.load_cb202()
        self.le201.setText("")
        self.le202.setText("")

    def add_cong_viec(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        mnv = self.lb004.text().strip()
        ten_cv = self.le207.text()

        if mnv and ten_cv:
            try:
                cursor.execute(
                    "INSERT INTO GHI_NHO_TEN_CONG_VIEC (MNV, Ten_cong_viec) VALUES (?, ?)",
                    (mnv, ten_cv)
                )
                connection.commit()
                self.msgbox("✅ Thêm công việc mới thành công")
            except Exception as e:
                print("Lỗi khi thêm công việc mới:", e)
                self.msgbox("⚠️ Tên công việc đã tồn tại hoặc có lỗi khác")
        else:
            self.msgbox("⚠️ Vui lòng nhập đầy đủ tên công việc mới")

        connection.close()
        self.load_cb207()
        self.le207.setText("")
    def add_chi_tiet_cong_viec(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        mnv = self.lb004.text().strip()
        ten_cv = self.le303.text()

        if mnv and ten_cv:
            try:
                cursor.execute(
                    "INSERT INTO GHI_NHO_TEN_CONG_VIEC (MNV, Ten_chi_tiet_cong_viec) VALUES (?, ?)",
                    (mnv, ten_cv)
                )
                connection.commit()
                self.msgbox("✅ Thêm chi tiết công việc mới thành công")
            except Exception as e:
                print("Lỗi khi thêm chi tiết công việc mới:", e)
                self.msgbox("⚠️ Tên chi tiết công việc đã tồn tại hoặc có lỗi khác")
        else:
            self.msgbox("⚠️ Vui lòng nhập đầy đủ tên hi tiết công việc mới")

        connection.close()
        self.load_cb302()
        self.le303.setText("")

    def add_chuc_danh(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_pb = self.le504.text().strip()
        ten_pb = self.le505.text().strip()
        nhom = self.le506.text().strip()
        chuc_danh = self.le507.text().strip()

        if ma_pb and ten_pb and chuc_danh:
            try:
                cursor.execute(
                    """INSERT INTO DANH_SACH_PHONG_BAN (Ma_phong_ban,Ten_phong_ban,Nhom,Chuc_danh) 
                    VALUES(?,?,?,?)""",
                    (ma_pb, ten_pb,nhom,chuc_danh)
                )
                connection.commit()
                self.msgbox("✅ Thêm chức danh công việc mới thành công")
            except Exception as e:
                self.msgbox(f"⚠️ Lỗi khi thêm chức danh công việc mới : {e}")
        else:
            self.msgbox("⚠️ Vui lòng nhập đầy đủ thông tin chức vụ công việc mới")

        connection.close()
        self.load_cb501()
        self.le504.setText("")
        self.le505.setText("")
        self.le506.setText("")
        self.le507.setText("")

    def load_cb200(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        nguoi_tao = self.lb004.text() + '-' + self.lb001.text()
        cursor.execute("SELECT ID_CV FROM DANH_SACH_CONG_VIEC WHERE Nguoi_cap_nhat = ?",(nguoi_tao,))
        results = cursor.fetchall()

        self.cb200.clear()  # Xóa các item hiện có
        for row in results:
            self.cb200.addItem(str(row[0])) 

        connection.close()
        self.cb200.setEditable(True)
        self.cb200.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb200.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb200.setStyleSheet("""
            QComboBox {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
            }
        """)
    def load_cb201(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_phong_ban = self.lb000.text()
        cursor.execute("SELECT DISTINCT NHOM FROM DANH_SACH_CBCNV WHERE Ma_phong_ban = ?",(ma_phong_ban,))
        results = cursor.fetchall()

        self.cb201.clear()  # Xóa các item hiện có
        for row in results:
            self.cb201.addItem(row[0])  # row[0] 

        connection.close()
    def change_cb200(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        id = self.cb200.currentText()
        nguoi_cap_nhat = self.lb004.text() + '-' + self.lb001.text()
        cursor.execute("SELECT * FROM DANH_SACH_CONG_VIEC WHERE ID_CV = ? and Nguoi_cap_nhat = ?",(id,nguoi_cap_nhat))
        results = cursor.fetchall()
 
        if results:
            ma_cong_ty = results[0][1]
            nhom = results[0][3]
            ma_du_an = results[0][4]
            ten_cong_viec = results[0][5]
            phan_loai_cv = results[0][6]
            ghi_chu = results[0][12]
            phan_loai_du_an = results[0][17]
            nhiem_vu = results[0][18]
            nhiem_vu_cu_the = results[0][19]
            chuc_nang = results[0][20]

            self.cb208.setCurrentIndex(self.cb208.findText(ma_cong_ty))
            self.cb201.setCurrentIndex(self.cb201.findText(nhom))
            self.cb202.setCurrentIndex(self.cb202.findText(ma_du_an))
            self.cb203.setCurrentIndex(self.cb203.findText(phan_loai_du_an))
            self.cb204.setCurrentIndex(self.cb204.findText(chuc_nang))
            self.cb205.setCurrentIndex(self.cb205.findText(nhiem_vu))
            self.cb206.setCurrentIndex(self.cb206.findText(nhiem_vu_cu_the))
            self.cb207.setCurrentIndex(self.cb207.findText(ten_cong_viec))
            self.cb210.setCurrentIndex(self.cb210.findText(phan_loai_cv))
            self.te201.setPlainText(ghi_chu)

        connection.close()
        self.show_tab_1
    def change_cb301(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        id = self.cb301.currentText()
        cursor.execute("""SELECT Chi_tiet_cong_viec,Ngay_bat_dau,Thoi_han,Thoi_luong,Tien_do,
                       Nguoi_thuc_hien,Ghi_chu,Diem_tien_do,Diem_chat_luong
                       FROM CHI_TIET_CONG_VIEC 
                       WHERE ID = ?""",(id,))
        results = cursor.fetchall()
 
        if results:
            chi_tiet_cong_viec = results[0][0]
            ngay_bat_dau = results[0][1]
            Thoi_han = results[0][2]
            Thoi_luong = results[0][3]
            Tien_do = int(results[0][4])
            Nguoi_thuc_hien = results[0][5]
            Ghi_chu = results[0][6]
            Diem_tien_do = int(results[0][7])
            Diem_chat_luong = int(results[0][8])

            self.cb302.setCurrentIndex(self.cb302.findText(chi_tiet_cong_viec))
            date_parts = ngay_bat_dau.split("-")
            year, month, day = map(int, date_parts)
            qdate = QDate(year, month, day)
            self.de301.setDate(qdate)
            date_parts = Thoi_han.split("-")
            year, month, day = map(int, date_parts)
            qdate = QDate(year, month, day)
            self.de302.setDate(qdate)
            self.le304.setText(str(Thoi_luong))
            self.le305.setText(str(Tien_do))
            self.le306.setText(str(Diem_tien_do))
            self.le307.setText(str(Diem_chat_luong))
            self.te302.setPlainText(Ghi_chu)

        connection.close()
    def change_cb500(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        mnv = self.cb500.currentText()
        cursor.execute("""SELECT MNV,Ho_ten,Ma_phong_ban,Ten_phong_ban,Nhom,Chuc_danh,SDT,Phan_quyen,Email
                       FROM DANH_SACH_CBCNV
                       WHERE MNV = ?""",(mnv,))
        results = cursor.fetchall()
 
        if results:
            mnv = results[0][0]
            ho_ten = results[0][1]
            ma_pb = results[0][2]
            ten_pb = results[0][3]
            nhom = results[0][4]
            chuc_danh = results[0][5]
            SDT = results[0][6]
            phan_quyen = results[0][7]
            email = results[0][8]

            self.le501.setText(mnv)
            self.le502.setText(ho_ten)
            self.cb501.setCurrentIndex(self.cb501.findText(ma_pb))
            self.lb501.setText(ten_pb)
            self.cb502.setCurrentIndex(self.cb502.findText(nhom))
            self.cb503.setCurrentIndex(self.cb503.findText(chuc_danh))
            self.le503.setText(SDT)
            self.cb504.setCurrentIndex(self.cb504.findText(phan_quyen))
            self.le508.setText(email)
        connection.close()

    def insert_cong_viec(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_cong_ty = self.cb208.currentText()
        ma_phong_ban = self.lb000.text()
        nhom = self.cb201.currentText()
        ma_du_an = self.cb202.currentText()
        ten_cong_viec = self.cb207.currentText()
        phan_loai_cv = self.cb210.currentText()
        ngay_bat_dau = date.today()
        thoi_luong = 0
        thoi_han = date.today()
        trang_thai = ""
        tien_do = 0
        ghi_chu = self.te201.toPlainText()
        diem_tien_do = 0
        diem_chat_luong = 0
        thoi_diem_cap_nhat = datetime.now()
        nguoi_cap_nhat = self.lb004.text() + '-' + self.lb001.text()
        phan_loai_du_an = self.cb203.currentText()
        nhiem_vu = self.cb205.currentText()
        nhiem_vu_cu_the = self.cb206.currentText()
        chuc_nang = self.cb204.currentText()
        try:
            cursor.execute("""INSERT INTO DANH_SACH_CONG_VIEC (Ma_cong_ty,Ma_phong_ban,Nhom,Ma_du_an,
                        Ten_cong_viec,Phan_loai_cv,Ngay_bat_dau,Thoi_luong,Thoi_han,Trang_thai,Tien_do,
                        Ghi_chu,Diem_tien_do,Diem_chat_luong,Thoi_diem_cap_nhat,Nguoi_cap_nhat,
                        Phan_loai_du_an,Nhiem_vu,Nhiem_vu_cu_the,Chuc_nang) 
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,(ma_cong_ty,ma_phong_ban,nhom,ma_du_an,ten_cong_viec,phan_loai_cv,ngay_bat_dau,
                                thoi_luong,thoi_han,trang_thai,tien_do,ghi_chu,diem_tien_do,diem_chat_luong,
                                thoi_diem_cap_nhat,nguoi_cap_nhat,phan_loai_du_an,nhiem_vu,nhiem_vu_cu_the,chuc_nang))
            connection.commit()
            self.msgbox("✅ Thêm công việc mới thành công")
            self.check_notifications()
            self.show_tab_1()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi thêm công việc mới: {e}")  

        connection.close()
    def insert_chi_tiet_cong_viec(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        id_cv = self.le301.text()
        nguoi_thuc_hien = self.cb304.currentText()
        chuc_danh = self.lb304.text()
        chi_tiet_cong_viec = self.cb302.currentText()
        ngay_bat_dau = self.de301.date().toString("yyyy-MM-dd")
        thoi_han = self.de302.date().toString("yyyy-MM-dd")
        try:
            thoi_luong = float(self.le304.text())
        except Exception  as e:
            self.msgbox("⚠️ Vui lòng nhập thời lượng (giờ)")
            return
        try:
            tien_do = int(self.le305.text())
        except Exception  as e:
            self.msgbox("⚠️ Vui lòng nhập tiến độ công việc")
            return
        trang_thai = "Chưa thực hiện" if tien_do == 0 else "Đã hoàn thành" if tien_do == 100 else "Đang thực hiện"
        ghi_chu = self.te302.toPlainText()
        diem_tien_do = 3
        diem_chat_luong = 3
        thoi_diem_cap_nhat = datetime.now()
        nguoi_tao = self.lb004.text() + '-' + self.lb001.text()
        diem_tien_do_x_thoi_luong = diem_tien_do * thoi_luong
        diem_chat_luong_x_thoi_luong = diem_chat_luong * thoi_luong
        tien_do_x_thoi_luong = tien_do * thoi_luong
        
        try:
            cursor.execute("""INSERT INTO CHI_TIET_CONG_VIEC (ID_CV,Nguoi_thuc_hien,Chuc_danh,
                        Chi_tiet_cong_viec,Ngay_bat_dau,Thoi_luong,Thoi_han,Trang_thai,Tien_do,
                        Ghi_chu,Diem_tien_do,Diem_chat_luong,Thoi_diem_cap_nhat,Nguoi_tao,
                        Diem_tien_do_x_thoi_luong,Diem_chat_luong_x_thoi_luong,tien_do_x_thoi_luong) 
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,(id_cv,nguoi_thuc_hien,chuc_danh,chi_tiet_cong_viec,ngay_bat_dau,
                                thoi_luong,thoi_han,trang_thai,tien_do,ghi_chu,diem_tien_do,diem_chat_luong,
                                thoi_diem_cap_nhat,nguoi_tao,diem_tien_do_x_thoi_luong,diem_chat_luong_x_thoi_luong,
                                tien_do_x_thoi_luong))
            connection.commit()
            self.msgbox("✅ Thêm chi tiết công việc mới thành công")
            self.search_chi_tiet_cong_viec()
            self.check_notifications()
            #send_email("user@example.com", "🔔 Nhắc việc", "Bạn có công việc mới vừa được tạo, vui lòng kiểm tra trên phần mềm nhật ký công việc!.")

        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi thêm chi tiết công việc mới: {e}")  

        connection.close()

    def save_cong_viec(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        id = self.cb200.currentText()
        ma_cong_ty = self.cb208.currentText()
        ma_phong_ban = self.lb000.text()
        nhom = self.cb201.currentText()
        ma_du_an = self.cb202.currentText()
        ten_cong_viec = self.cb207.currentText()
        phan_loai_cv = self.cb210.currentText()
        ghi_chu = self.te201.toPlainText()
        thoi_diem_cap_nhat = datetime.now()
        phan_loai_du_an = self.cb203.currentText()
        nhiem_vu = self.cb205.currentText()
        nhiem_vu_cu_the = self.cb206.currentText()
        chuc_nang = self.cb204.currentText()
        try:
            cursor.execute("""UPDATE DANH_SACH_CONG_VIEC SET 
                           Ma_cong_ty = ?,Nhom = ?,Ma_du_an = ?,
                        Ten_cong_viec = ?,Phan_loai_cv = ?,
                        Ghi_chu = ?,Thoi_diem_cap_nhat = ?,
                        Phan_loai_du_an = ?,Nhiem_vu = ?,Nhiem_vu_cu_the = ?,Chuc_nang = ?
                        WHERE ID_CV = ?
                        """,(ma_cong_ty,nhom,ma_du_an,ten_cong_viec,phan_loai_cv,
                                ghi_chu,thoi_diem_cap_nhat,phan_loai_du_an,nhiem_vu,nhiem_vu_cu_the,chuc_nang,id))
            connection.commit()
            self.msgbox("✅ Đã lưu thay đổi thông tin công việc")
            self.check_notifications()
            self.show_tab_1()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu thay đổi thông tin công việc: {e}")  

        connection.close()
    def save_chi_tiet_cong_viec(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        id = self.cb301.currentText()
        chi_tiet_cong_viec = self.cb302.currentText()
        ngay_bat_dau = self.de301.date().toString("yyyy-MM-dd")
        thoi_han = self.de302.date().toString("yyyy-MM-dd")
        thoi_luong = float(self.le304.text())
        tien_do = int(float(self.le305.text()))
        trang_thai = "Chưa thực hiện" if tien_do == 0 else "Đã hoàn thành" if tien_do == 100 else "Đang thực hiện"
        nguoi_thuc_hien = self.cb304.currentText()
        chuc_danh = self.lb304.text()
        ghi_chu = self.te302.toPlainText()
        thoi_diem_cap_nhat = datetime.now()
        diem_tien_do = int(float(self.le306.text()))
        diem_chat_luong = int(float(self.le307.text()))
        diem_tien_do_x_thoi_luong = diem_tien_do * thoi_luong
        diem_chat_luong_x_thoi_luong = diem_chat_luong * thoi_luong
        tien_do_x_thoi_luong = tien_do * thoi_luong
        try:
            cursor.execute("""UPDATE CHI_TIET_CONG_VIEC SET 
                           Chi_tiet_cong_viec = ?,Ngay_bat_dau = ?,Thoi_han = ?,
                        Thoi_luong = ?,Tien_do = ?,
                        Nguoi_thuc_hien = ?,Chuc_danh = ?,
                        Ghi_chu = ?,Thoi_diem_cap_nhat = ?,Diem_tien_do = ?,Diem_chat_luong = ?,
                        diem_tien_do_x_thoi_luong = ?, diem_chat_luong_x_thoi_luong = ?,
                        tien_do_x_thoi_luong = ?,trang_thai = ?
                        WHERE ID = ?
                        """,(chi_tiet_cong_viec,ngay_bat_dau,thoi_han,thoi_luong,tien_do,
                                nguoi_thuc_hien,chuc_danh,ghi_chu,thoi_diem_cap_nhat,diem_tien_do,diem_chat_luong,
                                diem_tien_do_x_thoi_luong,diem_chat_luong_x_thoi_luong,tien_do_x_thoi_luong,trang_thai,id))
            connection.commit()
            self.msgbox("✅ Đã lưu thay đổi thông tin chi tiết công việc")
            self.search_chi_tiet_cong_viec()
            self.check_notifications()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu thay đổi thông tin chi tiết công việc: {e}")  

        connection.close()
    def save_CBCNV(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        mnv = self.cb500.currentText()
        mnv_new = self.le501.text()
        ho_ten = self.le502.text()
        ma_pb = self.cb501.currentText()
        ten_pb = self.lb501.text()
        nhom = self.cb502.currentText()
        chuc_danh = self.cb503.currentText()
        sdt  = self.le503.text()
        phan_quyen = self.cb504.currentText()
        email = self.le508.text()
        try:
            cursor.execute("""UPDATE DANH_SACH_CBCNV SET 
                           MNV = ?,Ho_ten = ?,Ma_phong_ban = ?,
                        Ten_phong_ban = ?,Nhom = ?,
                        Chuc_danh = ?,SDT = ?,
                        Phan_quyen = ?,Email = ?
                        WHERE MNV = ?
                        """,(mnv_new,ho_ten,ma_pb,ten_pb,nhom,chuc_danh,sdt,phan_quyen,email,mnv))
            connection.commit()
            self.msgbox("✅ Đã lưu thay đổi thông tin CBCNV")
            self.search_CBCNV()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu thay đổi thông tin CBCNV: {e}")  

        connection.close()
    def insert_CBCNV(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        mnv_new = self.le501.text()
        ho_ten = self.le502.text()
        ma_pb = self.cb501.currentText()
        ten_pb = self.lb501.text()
        nhom = self.cb502.currentText()
        chuc_danh = self.cb503.currentText()
        sdt  = self.le503.text()
        phan_quyen = self.cb504.currentText()
        mat_khau = '123456'
        email = self.le508.text()
        if mnv_new and ho_ten:
            try:
                cursor.execute("""INSERT INTO DANH_SACH_CBCNV
                            (MNV,Ho_ten,Ma_phong_ban,
                            Ten_phong_ban,Nhom,
                            Chuc_danh,SDT,
                            Phan_quyen,Mat_khau,Email) 
                            VALUES(?,?,?,?,?,?,?,?,?,?)
                            """,(mnv_new,ho_ten,ma_pb,ten_pb,nhom,chuc_danh,sdt,phan_quyen,mat_khau,email))
                connection.commit()
                self.msgbox("✅ Đã thêm mới thông tin CBCNV")
                self.search_CBCNV()
            except Exception  as e:
                QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi thêm mới thông tin CBCNV: {e}")  
            self.search_CBCNV()
        else:
            self.msgbox("Mã nhân viên hoặc họ tên không được bỏ trống")
        connection.close()
    def delete_CBCNV(self):
        reply = QMessageBox.question(
        self,
        "Xác nhận xóa",
        "Bạn có chắc chắn muốn xóa nhân viên này không?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            connection = connect_to_db()
            cursor = connection.cursor()

            mnv = self.cb500.currentText()
            try:
                cursor.execute("""DELETE FROM DANH_SACH_CBCNV 
                            WHERE MNV = ?
                            """,(mnv,))
                connection.commit()

                self.msgbox("✅ Đã xóa thành công!")
            except Exception  as e:
                QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi xóa CBCNV: {e}")  

            connection.close()
            self.search_CBCNV()
    def delete_cong_viec(self):
        reply = QMessageBox.question(
        self,
        "Xác nhận xóa",
        "Khi bạn xóa công việc, sẽ xóa hết chi tiết công việc liên quan đến mã công việc này.\nBạn có chắc chắn muốn tiếp tục?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            connection = connect_to_db()
            cursor = connection.cursor()

            id = self.cb200.currentText()
            try:
                cursor.execute("""DELETE FROM DANH_SACH_CONG_VIEC 
                            WHERE ID_CV = ?
                            """,(id,))
                connection.commit()

                cursor.execute("""DELETE FROM CHI_TIET_CONG_VIEC 
                            WHERE ID_CV = ?
                            """,(id,))
                connection.commit()

                self.msgbox("✅ Đã xóa thành công!")
                self.show_tab_1()
            except Exception  as e:
                QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi xóa công việc: {e}")  

            connection.close()
            self.show_tab_1
            self.check_notifications()
    def delete_chi_tiet_cong_viec(self):
        reply = QMessageBox.question(
        self,
        "Xác nhận xóa",
        "Bạn có chắc chắn muốn xóa chi tiết công việc này?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            connection = connect_to_db()
            cursor = connection.cursor()

            id = self.cb301.currentText()
            try:
                cursor.execute("""DELETE FROM CHI_TIET_CONG_VIEC 
                            WHERE ID = ?
                            """,(id,))
                connection.commit()

                self.msgbox("✅ Đã xóa thành công!")
                self.search_chi_tiet_cong_viec()
            except Exception  as e:
                QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi xóa chi tiết công việc: {e}")  

            connection.close()
            self.search_chi_tiet_cong_viec()
            self.load_cb301()
            self.check_notifications()

    def load_cb204(self):
            connection = connect_to_db()
            cursor = connection.cursor()

            ma_phong_ban = self.lb000.text()
            phan_loai = self.cb203.currentText()
            cursor.execute("SELECT DISTINCT Ma_chuc_nang FROM CHUC_NANG_NHIEM_VU WHERE Ma_phong_ban = ? AND phan_loai = ?",(ma_phong_ban,phan_loai))
            results = cursor.fetchall()

            self.cb204.clear()  # Xóa các item hiện có
            for row in results:
                self.cb204.addItem(row[0])  # row[0] 

            connection.close()
    def load_cb205(self):
            connection = connect_to_db()
            cursor = connection.cursor()

            ma_pb = self.lb000.text()
            ma_chuc_nang = self.cb204.currentText()
            phan_quyen = self.lb005.text()
            chuc_danh = self.lb402.text()

            if (ma_chuc_nang == '4' and phan_quyen == "Quản lý") or ma_chuc_nang != '4':
                cursor.execute("""SELECT DISTINCT Ma_nhiem_vu FROM CHUC_NANG_NHIEM_VU 
                               WHERE Ma_chuc_nang = ? AND Ma_phong_ban = ?""",
                               (ma_chuc_nang,ma_pb))
                results = cursor.fetchall()
            else:
                cursor.execute("""SELECT DISTINCT Ma_nhiem_vu FROM CHUC_NANG_NHIEM_VU 
                               WHERE Ma_chuc_nang = ? and Chuc_nang = ? AND Ma_phong_ban = ?""",
                               (ma_chuc_nang,chuc_danh,ma_pb))
                results = cursor.fetchall()

            self.cb205.clear()  # Xóa các item hiện có
            for row in results:
                self.cb205.addItem(row[0])  # row[0] 

            connection.close()   
    def load_cb206(self):
            connection = connect_to_db()
            cursor = connection.cursor()

            ma_nhiem_vu = self.cb205.currentText()
            cursor.execute("SELECT DISTINCT Ma_nhiem_vu_cu_the FROM CHUC_NANG_NHIEM_VU WHERE Ma_nhiem_vu = ?",(ma_nhiem_vu,))
            results = cursor.fetchall()

            self.cb206.clear()  # Xóa các item hiện có
            for row in results:
                self.cb206.addItem(row[0])  # row[0] 

            connection.close() 
    def load_cb207(self):
            connection = connect_to_db()
            cursor = connection.cursor()

            mnv = self.lb004.text()
            cursor.execute("""SELECT Ten_cong_viec FROM GHI_NHO_TEN_CONG_VIEC WHERE  
                           Ten_cong_viec IS NOT NULL AND MNV = ?""",(mnv,))
            results = cursor.fetchall()

            self.cb207.clear()  # Xóa các item hiện có
            for row in results:
                self.cb207.addItem(row[0])  # row[0] 

            connection.close() 
            self.cb207.setEditable(True)
            self.cb207.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
            self.cb207.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
            self.cb207.setStyleSheet("""
                QComboBox {
                    background-color: #2e2e2e;
                    color: white;
                }
                QComboBox QAbstractItemView {
                    background-color: #3c3c3c;
                    color: white;
                }
            """)
    def load_cb202(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        phan_loai_du_an = self.cb203.currentText()
        if phan_loai_du_an == "Công việc hàng ngày":
            self.cb202.clear()
            self.lb202.setText("")
        else:
            cursor.execute("SELECT ma_du_an FROM danh_sach_du_an")
            results = cursor.fetchall()

            self.cb202.clear()  # Xóa các item hiện có
            for row in results:
                self.cb202.addItem(row[0])  # row[0] là 'ten_du_an'

        connection.close()
        self.cb202.setEditable(True)
        self.cb202.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb202.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb202.setStyleSheet("""
            QComboBox {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
            }
        """)
    def load_cb302(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        mnv = self.lb004.text()
        cursor.execute("""SELECT Ten_chi_tiet_cong_viec FROM GHI_NHO_TEN_CONG_VIEC WHERE  
                Ten_chi_tiet_cong_viec IS NOT NULL AND MNV = ?""",(mnv,))
        results = cursor.fetchall()

        self.cb302.clear()  # Xóa các item hiện có
        for row in results:
            self.cb302.addItem(row[0])  

        connection.close()
        self.cb302.setEditable(True)
        self.cb302.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb302.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb302.setStyleSheet("""
            QComboBox {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
                font-size: 13pt;              
            }
        """)
    def load_cb301(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        id_cv = self.le301.text()
        mnv = self.lb004.text() + '-' + self.lb001.text()
        phan_quyen = self.lb005.text()
        if phan_quyen == "Quản lý":
            cursor.execute("""SELECT ID FROM CHI_TIET_CONG_VIEC WHERE  
                ID_CV  = ?""",(id_cv,))
        else:
            cursor.execute("""SELECT ID FROM CHI_TIET_CONG_VIEC WHERE  
                ID_CV  = ? AND Nguoi_tao = ?""",(id_cv,mnv))
        results = cursor.fetchall()

        self.cb301.clear()  # Xóa các item hiện có
        for row in results:
            self.cb301.addItem(str(row[0]))

        connection.close()
        self.cb301.setEditable(True)
        self.cb301.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb301.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb301.setStyleSheet("""
            QComboBox {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
                font-size: 13pt;              
            }
        """)
    def load_cb304(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_pb = self.lb000.text()
        cursor.execute("""SELECT MNV, Ho_ten FROM DANH_SACH_CBCNV WHERE  
                Ma_phong_ban  = ?""",(ma_pb,))
        results = cursor.fetchall()

        self.cb304.clear()  # Xóa các item hiện có
        for row in results:
            self.cb304.addItem(str(row[0]) + '-' + str(row[1]))

        connection.close()
        self.cb304.setEditable(True)
        self.cb304.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb304.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb304.setStyleSheet("""
            QComboBox {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
                font-size: 13pt;              
            }
        """)
    def load_cb500(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        cursor.execute("""SELECT MNV FROM DANH_SACH_CBCNV""")
        results = cursor.fetchall()

        self.cb500.clear()  # Xóa các item hiện có
        for row in results:
            self.cb500.addItem(str(row[0]))

        connection.close()
        self.cb500.setEditable(True)
        self.cb500.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb500.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb500.setStyleSheet("""
            QCombo0Box {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
                font-size: 13pt;              
            }
        """)
    def load_cb501(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        cursor.execute("""SELECT DISTINCT MA_PHONG_BAN FROM DANH_SACH_PHONG_BAN""")
        results = cursor.fetchall()

        self.cb501.clear()  # Xóa các item hiện có
        for row in results:
            self.cb501.addItem(str(row[0]))

        connection.close()
        self.cb501.setEditable(True)
        self.cb501.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb501.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb501.setStyleSheet("""
            QCombo0Box {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
                font-size: 13pt;              
            }
        """)
    def load_cb502(self):
        connection = connect_to_db()
        cursor = connection.cursor()
        ma_pb = self.cb501.currentText()
        cursor.execute("""SELECT DISTINCT NHOM FROM DANH_SACH_PHONG_BAN WHERE MA_PHONG_BAN = ?""",(ma_pb,))
        results = cursor.fetchall()

        self.cb502.clear()  # Xóa các item hiện có
        for row in results:
            self.cb502.addItem(str(row[0]))

        connection.close()
        self.cb502.setEditable(True)
        self.cb502.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb502.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb502.setStyleSheet("""
            QCombo0Box {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
                font-size: 13pt;              
            }
        """)
    def load_cb503(self):
        connection = connect_to_db()
        cursor = connection.cursor()
        ma_pb = self.cb501.currentText()
        nhom = self.cb502.currentText()
        cursor.execute("""SELECT DISTINCT chuc_danh FROM DANH_SACH_PHONG_BAN WHERE MA_PHONG_BAN = ? and nhom = ?""",(ma_pb,nhom))
        results = cursor.fetchall()

        self.cb503.clear()  # Xóa các item hiện có
        for row in results:
            self.cb503.addItem(str(row[0]))

        connection.close()
        self.cb503.setEditable(True)
        self.cb503.completer().setFilterMode(QtCore.Qt.MatchContains)  # Gợi ý theo từ khóa chứa
        self.cb503.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.cb503.setStyleSheet("""
            QCombo0Box {
                background-color: #2e2e2e;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: white;
                font-size: 13pt;              
            }
        """)

    def load_lb202(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_du_an = self.cb202.currentText()
        cursor.execute("SELECT ten_du_an FROM danh_sach_du_an where ma_du_an = ?",(ma_du_an,))
        results = cursor.fetchone()
        if results:
            self.lb202.setText(results[0])
        else:
            self.lb202.setText("")

    def load_lb304(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        mnv = self.cb304.currentText().split("-")[0].strip()
        cursor.execute("SELECT Chuc_danh FROM DANH_SACH_cbcnv where MNV = ?",(mnv,))
        results = cursor.fetchone()
        if results:
            self.lb304.setText(results[0])
        else:
            self.lb304.setText("")
    def load_lb204(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_chuc_nang = self.cb204.currentText()
        if ma_chuc_nang == '4':
            self.lb204.setText("Công việc chuyên môn theo vị trí")
        else:
            cursor.execute("SELECT Chuc_nang FROM CHUC_NANG_NHIEM_VU where Ma_chuc_nang = ?",(ma_chuc_nang,))
            results = cursor.fetchone()
            if results:
                self.lb204.setText(results[0])
            else:
                self.lb204.setText("")
    def load_lb205(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_nhiem_vu = self.cb205.currentText()
        cursor.execute("SELECT Nhiem_vu FROM CHUC_NANG_NHIEM_VU where Ma_nhiem_vu = ?",(ma_nhiem_vu,))
        results = cursor.fetchone()
        if results:
            self.lb205.setText(results[0])
        else:
            self.lb205.setText("")
    def load_lb206(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_nhiem_vu_cu_the = self.cb206.currentText()
        cursor.execute("SELECT Nhiem_vu_cu_the FROM CHUC_NANG_NHIEM_VU where Ma_nhiem_vu_cu_the = ?",(ma_nhiem_vu_cu_the,))
        results = cursor.fetchone()
        if results:
            self.lb206.setText(results[0])
        else:
            self.lb206.setText("")
    def load_lb501(self):
        connection = connect_to_db()
        cursor = connection.cursor()

        ma_pb = self.cb501.currentText()
        cursor.execute("SELECT Ten_phong_ban FROM DANH_SACH_PHONG_BAN where Ma_phong_ban = ?",(ma_pb,))
        results = cursor.fetchone()
        if results:
            self.lb501.setText(results[0])
        else:
            self.lb501.setText("")
    def search_cong_viec(self):
        ma_cong_ty = self.cb103.currentText()
        if ma_cong_ty == "Tất cả":
            ma_cong_ty = ""  

        ma_phong_ban = self.cb104.currentText()
        if ma_phong_ban == "Tất cả":
            ma_phong_ban = ""  

        tu_ngay = self.de101.date().toString("yyyy-MM-dd")
        den_ngay = self.de102.date().toString("yyyy-MM-dd")

        trang_thai = self.cb101.currentText()
        if trang_thai == "Tất cả":
            trang_thai = ""

        phan_loai_du_an = self.cb105.currentText()
        if phan_loai_du_an == "Tất cả":
            phan_loai_du_an = ""
        
        phan_loai_cong_viec = self.cb102.currentText()
        if phan_loai_cong_viec == "Tất cả":
            phan_loai_cong_viec = ""
       
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        sql = f"""
                SELECT ID_CV,Ma_cong_ty,Ma_phong_ban,Nhom,Phan_loai_du_an,Ma_du_an,Chuc_nang,
                Nhiem_vu,Nhiem_vu_cu_the,Ten_cong_viec,Phan_loai_cv,Ngay_bat_dau,Thoi_luong,
                Thoi_han,Trang_thai,Tien_do,Ghi_chu,Diem_tien_do,Diem_chat_luong,Thoi_diem_cap_nhat,
                Nguoi_cap_nhat
                FROM DANH_SACH_CONG_VIEC WHERE 
                Trang_thai LIKE ? and Ma_cong_ty LIKE ? and Phan_loai_du_an LIKE ? and
                Phan_loai_cv LIKE ? and Ma_phong_ban LIKE ? 
                and (Ngay_bat_dau BETWEEN ? AND ?)
                ORDER BY ID_CV DESC
            """
        cursor.execute(sql,(f"%{trang_thai}%",f"%{ma_cong_ty}%",f"%{phan_loai_du_an}%",
                            f"%{phan_loai_cong_viec}%",f"%{ma_phong_ban}%",tu_ngay,den_ngay))
        results = cursor.fetchall()
        # Xóa dữ liệu cũ trong TableWidget
        self.tableWidget.setRowCount(0)
        
        # Hiển thị dữ liệu trong TableWidget
        for row_idx, row_data in enumerate(results):
            self.tableWidget.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                if col_idx == 15:
                    value = round(value, 0)
                if col_idx == 17:
                    value = round(value, 2)
                if col_idx == 18:
                    value = round(value, 2)
                if col_idx == 19:
                    value = value[:16]
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.tableWidget.setItem(row_idx, col_idx, item)
 
        self.tableWidget.resizeColumnsToContents()

        # Gọi hàm tổng số dòng
        self.tong_so_dong_tab_1

    def search_chi_tiet_cong_viec(self):
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        id_cv = self.le301.text()
        sql = f"""
                SELECT ID,Chi_tiet_cong_viec,Thoi_luong,Ngay_bat_dau,
                Thoi_han,Tien_do,Trang_thai,Ghi_chu,Nguoi_thuc_hien,Chuc_danh,Diem_tien_do,Diem_chat_luong,
                Thoi_diem_cap_nhat,Nguoi_tao
                FROM CHI_TIET_CONG_VIEC WHERE ID_CV = ?
                ORDER BY ID DESC
            """
        cursor.execute(sql,(id_cv,))
        results = cursor.fetchall()
        # Xóa dữ liệu cũ trong TableWidget
        self.tableWidget_2.setRowCount(0)
        
        # Hiển thị dữ liệu trong TableWidget
        for row_idx, row_data in enumerate(results):
            self.tableWidget_2.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                if col_idx == 12:
                    value = value[:16]
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.tableWidget_2.setItem(row_idx, col_idx, item)
        
        self.tableWidget_2.resizeColumnsToContents()

    def search_CBCNV(self):
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        sql = f"""
                SELECT MNV,Ho_ten,Ma_phong_ban,Ten_phong_ban,Nhom,Chuc_danh,SDT,Phan_quyen,Email
                FROM DANH_SACH_CBCNV
                ORDER BY MNV
            """
        cursor.execute(sql,)
        results = cursor.fetchall()
        # Xóa dữ liệu cũ trong TableWidget
        self.tableWidget_3.setRowCount(0)
        
        # Hiển thị dữ liệu trong TableWidget
        for row_idx, row_data in enumerate(results):
            self.tableWidget_3.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.tableWidget_3.setItem(row_idx, col_idx, item)
        
        self.tableWidget_3.resizeColumnsToContents()

    def show_login_tab(self):
        self.tabWidget.setCurrentIndex(0) 
        self.tb001.setText("")
        self.tb002.setText("")
        self.lb000.setText("")
        self.lb001.setText("")
        self.lb002.setText("")
        self.lb003.setText("Nhật ký công việc")
        self.menuBar.setVisible(False)
                        
    def show_tab_1(self):
        self.tabWidget.setCurrentIndex(1) 
        self.lb003.setText("Danh sách công việc")
        self.search_cong_viec()
     
    def show_tab_2(self):
        self.menuBar.setVisible(True)
        self.tabWidget.setCurrentIndex(2) 
        self.lb003.setText("Thêm công việc mới")  
        self.load_cb201()
        self.load_cb202()
        self.load_cb204()
        self.load_cb207()
        self.rd201.setChecked(True)

    def show_tab_3(self,id_cv,ma_pb,ten_cv,ghi_chu):
            self.tabWidget.setCurrentIndex(3) 
            self.lb003.setText("Chi tiết công việc")
            self.le301.setText(id_cv)
            self.le301.setReadOnly(True)
            self.le302.setText(ten_cv)
            self.le302.setReadOnly(True)
            self.te301.setPlainText(ghi_chu)
            self.te301.setReadOnly(True)
            if ma_pb == self.lb000.text():
                self.frame_301.setVisible(True)
                self.rd301.setChecked(True)
            else:
                self.frame_301.setVisible(False)
            self.load_cb302()
            self.load_cb304()
            self.search_chi_tiet_cong_viec()
    def show_tab_4(self):
        self.tabWidget.setCurrentIndex(4) 
        self.lb003.setText("Đổi mật khẩu")
    def show_tab_5(self):
        self.tabWidget.setCurrentIndex(5) 
        self.lb003.setText("Danh sách CBCNV")
        ma_pb = self.lb000.text()
        if ma_pb == 'HR':
            self.frame_500.setVisible(True)
        else:
            self.frame_500.setVisible(False)
        self.search_CBCNV()
        self.rd501.setChecked(True)
        self.load_cb501()
   
    def on_table_double_click(self, item):
        row = item.row()
        id_item = self.tableWidget.item(row, 0)
        ma_pb_item = self.tableWidget.item(row, 2)
        ten_cv_item = self.tableWidget.item(row, 9)
        ghi_chu_item = self.tableWidget.item(row, 16)
        if id_item:
            id_cv = id_item.text()
            ma_pb = ma_pb_item.text()
            ten_cv = ten_cv_item.text()
            ghi_chu = ghi_chu_item.text()

            self.show_tab_3(id_cv,ma_pb,ten_cv,ghi_chu) 
        
    def tong_so_dong_tab_1(self):
        rows = self.tableWidget.rowCount()
        self.lb101.setText(f"Tổng số dòng dữ liệu: {rows}")
        self.lb101.setStyleSheet("color: rgb(0, 255, 0);")
              
    def delete_selected_rows_QA(self):
        # Lấy ID danh sách các hàng được chọn
        selected_IDs  = set(self.tableWidget.item(index.row(),31).text() 
                          for index in self.tableWidget.selectedIndexes()
                          if self.tableWidget.item(index.row(),31) is not None)
        # Kiểm tra nếu không có hàng nào được chọn
        if not selected_IDs:  # Tập hợp rỗng
            QMessageBox.information(self, "Thông báo", "Chưa có dòng nào được chọn")
            return
        # Hiển thị cảnh báo xác nhận
        reply = QMessageBox.question(
            self,
            "Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa {len(selected_IDs)} dòng đã chọn không?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.No:
            return  # Không làm gì nếu người dùng chọn "No"
        try:
            connection = connect_to_db()
            if connection is None:
                self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
                return
            cursor = connection.cursor()
            # Chuyển danh sách ID thành chuỗi tham số
            placeholders = ", ".join(["?"] * len(selected_IDs))
            query = f"DELETE FROM QA_PRT WHERE ID IN ({placeholders})"
            # Thực thi câu lệnh với tham số
            cursor.execute(query, tuple(selected_IDs))
            connection.commit()
            
            QMessageBox.information(self, "Thông báo", f"Xóa thành công {len(selected_IDs)} dữ liệu!")
            self.search_QA()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi xóa dữ liệu: {e}")    
        finally:
            if connection:
                connection.close()   

    def delete_selected_rows_Checker(self):
        # Lấy ID danh sách các hàng được chọn
        selected_IDs  = set(self.tableWidget_2.item(index.row(),31).text() 
                          for index in self.tableWidget_2.selectedIndexes()
                          if self.tableWidget_2.item(index.row(),31) is not None)
        # Kiểm tra nếu không có hàng nào được chọn
        if not selected_IDs:  # Tập hợp rỗng
            QMessageBox.information(self, "Thông báo", "Chưa có dòng nào được chọn")
            return
        # Hiển thị cảnh báo xác nhận
        reply = QMessageBox.question(
            self,
            "Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa {len(selected_IDs)} dòng đã chọn không?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.No:
            return  # Không làm gì nếu người dùng chọn "No"
        try:
            connection = connect_to_db()
            if connection is None:
                self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
                return
            cursor = connection.cursor()
            # Chuyển danh sách ID thành chuỗi tham số
            placeholders = ", ".join(["?"] * len(selected_IDs))
            query = f"DELETE FROM CHECKER_PRT WHERE ID IN ({placeholders})"
            # Thực thi câu lệnh với tham số
            cursor.execute(query, tuple(selected_IDs))
            connection.commit()
            
            QMessageBox.information(self, "Thông báo", f"Xóa thành công {len(selected_IDs)} dữ liệu!")
            self.search_Checker()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi xóa dữ liệu: {e}")    
        finally:
            if connection:
                connection.close()   

    def tai_xuong_file_excel(self):
        connection = connect_to_db()
        cursor = connection.cursor()
        sql = """
            SELECT ID_CV,Chi_tiet_cong_viec,Thoi_luong,Ngay_bat_dau,
                Thoi_han,Tien_do,Trang_thai,Ghi_chu,Nguoi_thuc_hien,
                Chuc_danh,Diem_tien_do,Diem_chat_luong,Thoi_diem_cap_nhat,Nguoi_tao
            FROM CHI_TIET_CONG_VIEC
            ORDER BY ID DESC
        """
        cursor.execute(sql)
        results = cursor.fetchall()
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Chi tiết công việc"

        # Tạo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Chi tiết công việc"

        # Tên cột
        headers = [
            "ID công việc", "Chi tiết công việc", "Thời lượng", "Ngày bắt đầu",
            "Thời hạn", "Tiến độ", "Trạng thái", "Ghi chú", "Người thực hiện",
            "Chức danh", "Điểm tiến độ", "Điểm chất lượng", "Thời điểm cập nhật", "Người tạo"
        ]
        ws.append(headers)

        # Ghi dữ liệu
        for row in results:
            row = list(row)
            # Cắt thời điểm cập nhật (nếu có giá trị) để lấy 16 ký tự đầu (yyyy-mm-dd hh:mm)
            if row[12]:
                row[12] = str(row[12])[:16]
            ws.append(row)

        # Tự động điều chỉnh độ rộng cột
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Open file dialog to select save location
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Tải xuống file", 
            f"Chi tiết công việc.xlsx", 
            "Excel Files (*.xlsx);;All Files (*)", 
            options=options
        )

        if not file_path:
            # QMessageBox.information(self, "Thông báo", "Bạn đã hủy việc tải xuống.")
            return

        # Save the file to the selected location
        try:
            wb.save(file_path)
            QMessageBox.information(self, "Thông báo", f"File đã được tải xuống thành công tại:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu file: {e}")

    def tai_xuong_danh_sach_cong_viec(self):
        connection = connect_to_db()
        cursor = connection.cursor()
        sql = """
            SELECT ID_CV,Ma_cong_ty,Ma_phong_ban,Nhom,Ma_du_an,Ten_cong_viec,Phan_loai_cv,Ngay_bat_dau,
                Thoi_luong,Thoi_han,Tien_do,Trang_thai,Ghi_chu,Diem_tien_do,Diem_chat_luong,
                Phan_loai_du_an,Chuc_nang,Nhiem_vu,Nhiem_vu_cu_the,
                Thoi_diem_cap_nhat,Nguoi_cap_nhat
            FROM DANH_SACH_CONG_VIEC
            ORDER BY ID_CV DESC
        """
        cursor.execute(sql)
        results = cursor.fetchall()
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Danh sách công việc"

        # Tạo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách công việc"

        # Tên cột
        headers = [
            "ID công việc","Mã phòng ban", "Mã công ty","Nhóm","Mã dự án","Tên công việc","Phân loại công việc",
            "Ngày bắt đầu","Thời lượng", "Thời hạn","Tiến độ", "Trạng thái", "Ghi chú", 
            "Điểm tiến độ", "Điểm chất lượng","Phân loại dự án","Chức năng","Nhiệm vụ",
            "Nhiệm vụ cụ thể", "Thời điểm cập nhật", "Người tạo"
        ]
        ws.append(headers)

        # Ghi dữ liệu
        for row in results:
            row = list(row)
            # Cắt thời điểm cập nhật (nếu có giá trị) để lấy 16 ký tự đầu (yyyy-mm-dd hh:mm)
            if row[19]:
                row[19] = str(row[19])[:16]
            ws.append(row)

        # Tự động điều chỉnh độ rộng cột
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Open file dialog to select save location
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Tải xuống file", 
            f"Danh sách công việc.xlsx", 
            "Excel Files (*.xlsx);;All Files (*)", 
            options=options
        )

        if not file_path:
            # QMessageBox.information(self, "Thông báo", "Bạn đã hủy việc tải xuống.")
            return

        # Save the file to the selected location
        try:
            wb.save(file_path)
            QMessageBox.information(self, "Thông báo", f"File đã được tải xuống thành công tại:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu file: {e}")
           
def main():
    app = QApplication(sys.argv)
    window = MainApp()
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    window.show()
    app.exec_()

if __name__ == '__main__':
    main()

